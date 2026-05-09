#!/usr/bin/env python3
"""ts_clean.py — Time Series Data Cleaning for Time Series Analysis Skill"""

import zipfile, xml.etree.ElementTree as ET, os, argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import Counter

# ----- 读取 xlsx -----

def read_raw_data(filepath):
    with zipfile.ZipFile(filepath) as z:
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            ss_root = ET.parse(z.open('xl/sharedStrings.xml')).getroot()
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            for si in ss_root.findall('s:si', ns):
                t = si.find('s:t', ns)
                shared_strings.append(t.text if t is not None else '')
        root = ET.parse(z.open('xl/worksheets/sheet1.xml')).getroot()
        ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        result = []
        for r in root.findall('.//s:row', ns):
            rn = int(r.get('r', 0))
            rd = {}
            for c in r.findall('s:c', ns):
                col = ''.join(ch for ch in c.get('r', '') if ch.isalpha())
                ct = c.get('t', '')
                v = c.find('s:v', ns)
                rv = v.text if v is not None else ''
                if ct == 's' and rv.isdigit() and int(rv) < len(shared_strings):
                    rd[col] = shared_strings[int(rv)]
                else:
                    rd[col] = rv
            result.append((rn, rd))
    return result

# ----- 非样本行判断 -----

def is_non_sample(row_data):
    if not row_data or all(v == '' or v is None for v in row_data.values()):
        return True
    cols = list(row_data.keys())
    if cols and (row_data.get(cols[0], '') == '' or row_data.get(cols[0]) is None):
        non_empty = [v for v in row_data.values() if v and str(v).strip()]
        if non_empty:
            return True
    return False

# ----- 时间格式处理 -----

def excel_serial_to_datetime(s):
    if not s: return None
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(s))
    except (ValueError, TypeError):
        return None

def parse_date_string(s):
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d', '%Y-%m', '%Y/%m',
                '%d/%m/%Y', '%m/%d/%Y']:
        try:
            return datetime.strptime(s.strip(), fmt)
        except:
            continue
    return None

def detect_time_format(value):
    if not value: return None
    v = value.strip()
    if v.replace('.', '').replace('-', '').isdigit():
        try:
            if 1 < float(v) < 80000:
                return 'serial'
        except:
            pass
    if parse_date_string(v):
        return 'string'
    return None

# ----- 季度映射 -----

def month_to_quarter(m):
    if 1 <= m <= 3: return 1
    if 4 <= m <= 6: return 2
    if 7 <= m <= 9: return 3
    return 4

# ----- 小数精度自动检测 -----

def detect_majority_decimal_places(raw_values, threshold=0.5):
    dc, ni = Counter(), 0
    for val in raw_values:
        if not val: continue
        try: f = float(val)
        except: continue
        if f == int(f): continue
        ni += 1
        n = len(val.strip().split('.')[1])
        dc[n] += 1
    if ni == 0: return 2
    for n, cnt in dc.items():
        if cnt / ni > threshold: return n
    return dc.most_common(1)[0][0] if dc else 2

# ----- 主清洗函数 -----

def clean_ts_data(filepath, time_col='A', value_cols=None, value_names=None,
                  time_format='auto', frequency='daily', decimal_places=None,
                  output_xlsx=None, output_txt=None):
    if value_cols is None: value_cols = ['B']
    if value_names is None: value_names = ['数值']

    time_vars_map = {
        'daily': ['年份', '季度', '月份', '日'],
        'monthly': ['年份', '季度', '月份'],
        'quarterly': ['年份', '季度'],
        'yearly': ['年份'],
    }
    time_vars = time_vars_map.get(frequency, ['年份'])
    sort_keys_map = {
        'daily': ['year', 'quarter', 'month', 'day'],
        'monthly': ['year', 'month'],
        'quarterly': ['year', 'quarter'],
        'yearly': ['year'],
    }
    sort_keys = sort_keys_map.get(frequency, ['year'])

    print(f'读取原始文件: {filepath}')
    raw = read_raw_data(filepath)
    print(f'  总行数: {len(raw)}')

    sample_rows, skipped = [], 0
    for rn, rd in raw:
        if is_non_sample(rd):
            skipped += 1
            continue
        if not rd.get(time_col, ''):
            skipped += 1
            continue
        if not all(rd.get(vc, '') for vc in value_cols):
            skipped += 1
            continue
        sample_rows.append((rn, rd))
    print(f'  保留: {len(sample_rows)} 行, 跳过: {skipped} 行')

    if len(sample_rows) == 0:
        print('[错误] 没有找到有效样本行！')
        return

    # 小数精度
    if decimal_places is None:
        all_vals = [rd.get(vc, '') for _, rd in sample_rows for vc in value_cols if rd.get(vc, '')]
        decimal_places = detect_majority_decimal_places(all_vals)
        print(f'  -> 自动检测小数精度: {decimal_places} 位')

    # 时间格式
    if time_format == 'auto':
        first_val = sample_rows[0][1].get(time_col, '')
        detected = detect_time_format(first_val)
        time_format = detected if detected else 'serial'
        print(f'  -> 自动检测时间格式: {time_format}')

    cleaned = []
    time_errors = 0
    for rn, rd in sample_rows:
        tv = rd.get(time_col, '')
        dt = None
        if time_format == 'serial':
            dt = excel_serial_to_datetime(tv)
        elif time_format == 'string':
            dt = parse_date_string(tv)
        if dt is None:
            time_errors += 1
            continue
        rec = {'year': dt.year, 'quarter': month_to_quarter(dt.month),
               'month': dt.month, 'day': dt.day}
        ok = True
        for i, vc in enumerate(value_cols):
            try:
                rec[f'val_{i}'] = round(float(rd.get(vc, 0)), decimal_places)
            except:
                ok = False
                break
        if ok:
            cleaned.append(rec)

    cleaned.sort(key=lambda x: tuple(x[k] for k in sort_keys))

    first, last = cleaned[0], cleaned[-1]
    print(f'  有效数据: {len(cleaned)} 条')
    if frequency == 'daily':
        print(f'  时间范围: {first["year"]}-{first["month"]:02d}-{first["day"]:02d} ~ {last["year"]}-{last["month"]:02d}-{last["day"]:02d}')
    elif frequency == 'monthly':
        print(f'  时间范围: {first["year"]}-{first["month"]:02d} ~ {last["year"]}-{last["month"]:02d}')
    elif frequency == 'quarterly':
        print(f'  时间范围: {first["year"]}Q{first["quarter"]} ~ {last["year"]}Q{last["quarter"]}')
    else:
        print(f'  时间范围: {first["year"]} ~ {last["year"]}')

    # 导出 xlsx
    all_vars = time_vars + value_names
    wb = Workbook()
    ws = wb.active
    ws.title = '清洗后数据'
    hf = Font(bold=True, size=11)
    ha = Alignment(horizontal='center')
    key_map = {'年份': 'year', '季度': 'quarter', '月份': 'month', '日': 'day'}
    for ci, v in enumerate(all_vars, 1):
        c = ws.cell(1, ci, v)
        c.font = hf; c.alignment = ha
    for ri, rec in enumerate(cleaned, 2):
        for vi, tv in enumerate(time_vars):
            ws.cell(ri, vi + 1, rec.get(key_map.get(tv, ''), ''))
        for i in range(len(value_names)):
            ws.cell(ri, len(time_vars) + i + 1, rec.get(f'val_{i}', ''))
    for ci in range(1, len(all_vars) + 1):
        ws.column_dimensions[chr(64 + ci)].width = 16
    if output_xlsx:
        wb.save(output_xlsx)
        print(f'  {output_xlsx}')

    # 导出清洗流程
    if output_txt:
        lines = []
        lines.append('时间序列数据清洗流程')
        lines.append('')
        lines.append(f'原始文件: {filepath}')
        lines.append(f'有效观测: {len(cleaned)} 条')
        lines.append(f'小数精度: {decimal_places} 位')
        lines.append(f'时间范围: ...')
        with open(output_txt, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        print(f'  {output_txt}')

    print(f'清洗完成！共 {len(cleaned)} 条有效数据')

# ----- CLI -----

def main():
    parser = argparse.ArgumentParser(description='时间序列数据清洗')
    parser.add_argument('--input', required=True)
    parser.add_argument('--time-col', default='A')
    parser.add_argument('--value-cols', default='B')
    parser.add_argument('--value-names', default=None)
    parser.add_argument('--time-format', default='auto', choices=['auto', 'serial', 'string'])
    parser.add_argument('--frequency', default='daily', choices=['daily', 'monthly', 'quarterly', 'yearly'])
    parser.add_argument('--decimal-places', type=int, default=None)
    parser.add_argument('--output-xlsx', default=None)
    parser.add_argument('--output-txt', default=None)
    parser.add_argument('--probe', action='store_true')
    args = parser.parse_args()

    value_cols = [c.strip() for c in args.value_cols.split(',')]
    value_names = [n.strip() for n in args.value_names.split(',')] if args.value_names else [f'列{c}' for c in value_cols]
    if len(value_names) < len(value_cols):
        value_names += [f'列{c}' for c in value_cols[len(value_names):]]

    base = os.path.splitext(args.input)[0]
    if args.output_xlsx is None:
        args.output_xlsx = f'{base}_清洗后.xlsx'
    if args.output_txt is None:
        args.output_txt = f'{base}_清洗流程.txt'

    if args.probe:
        print(f'探查文件: {args.input}')
        return

    clean_ts_data(args.input, args.time_col, value_cols, value_names,
                  args.time_format, args.frequency, args.decimal_places,
                  args.output_xlsx, args.output_txt)

if __name__ == '__main__':
    main()
